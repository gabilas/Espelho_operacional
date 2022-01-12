from openpyxl import Workbook, load_workbook
import time

def main():

    #Login de Usuário
    usuario = str(input("Qual o seu usuário?\n")).lower()
    mes = str(input("Qual o mês?\n")).upper()
    ano = str(input("Qual o ano?\n")).upper()
    periodo = str(input("Qual o periodo?\n'01a15' ou '16a31'?\n")).lower()

    #Coletar informações dos colaboradores na base de dados
    Planilha_base_de_dados = load_workbook("C:\\Users\\{}\\Documents\\Colaboradores\\Colaboradores.xlsx".format(usuario))
    Aba = Planilha_base_de_dados.active
    #Planilha de Justificativa
    Planilha_Espelho = load_workbook("C:\\Users\\{}\\Documents\\Espelho de Ponto\\OCORRENCIAS.xlsx".format(usuario))
    Aba_ESP = Planilha_Espelho.active
    #Txt com a justificativa
    caminho_saida =open("C:\\Users\\{}\\Documents\\Espelho de Ponto\\{}\\{}{} - {}\\Saida.txt".format(usuario, ano, mes, ano, periodo), 'w') # Saída do resultado

    funcionários = []
    matriculas = []
    funções = []
    equipes= []

    
    for celula_nome in Aba['B']:  
        linha_nome = celula_nome.row
        nome = str(Aba["B{}".format(linha_nome)].value)
        if nome == "Funcionário":
            time.sleep(0.00001)
        else:
            funcionários.append(nome)

    for celula_matricula in Aba['A']:  
        linha_matricula = celula_matricula.row
        matricula = str(Aba["A{}".format(linha_matricula)].value)
        if matricula == "Matricula":
            time.sleep(0.00001)
        else:
            matriculas.append(matricula)

    for celula_função in Aba['C']:  
        linha_função= celula_função.row
        função = str(Aba["C{}".format(linha_função)].value)
        if função == "Função":
            time.sleep(0.00001)
        else:
            funções.append(função)

    for celula_equipe in Aba['D']:  

        linha_equipe = celula_equipe.row
        equipe = str(Aba["D{}".format(linha_equipe)].value)
        if equipe == "Equipe":
            time.sleep(0.00001)
        else:
          equipes.append(equipe)

    #caminho_saida.write('- Registro do Ponto Manual:\n\n')

    ultima_matricula = ""
    
    for celula in Aba_ESP['D']:
        linha = celula.row
        mat = str(Aba_ESP['D{}'.format(linha)].value)
        if mat in matriculas:
            nome = str(Aba_ESP['E{}'.format(linha)].value)
            Aba_ESP['S{}'.format(linha)] = nome
            Aba_ESP['T{}'.format(linha)] = equipes[matriculas.index(mat)]

            #caminho_saida.write('{} - Mat.{}\n'.format(nome,matriculas[matriculas.index(mat)]))

            data = str(Aba_ESP['L{}'.format(linha)].value)
            motivo = str(Aba_ESP['P{}'.format(linha)].value)
            if motivo == "ESQUECEU DE REGISTRAR":
                #if mat != ultima_matricula:
                    #caminho_saida.write('{} - Mat.{}\n'.format(nome,matriculas[matriculas.index(mat)]))
                    ultima_matricula = mat
                    #Aba_ESP['Q{}'.format(linha)] = "NATAL"
                    caminho_saida.write('ESQUCEU DE REGISTRAR: {} - Data:{}\n'.format(nome,data))
            if motivo == "NÃO CADASTRADO":
                    caminho_saida.write('NÃO CADASTRADO: {} - Data:{}\n'.format(nome,data))
                      
        else:
            time.sleep(0.0001)
    
    Planilha_Espelho.save("C:\\Users\\{}\\Documents\\Espelho de Ponto\\OCORRENCIAS.xlsx".format(usuario))
    caminho_saida.close()

main()